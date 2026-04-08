from __future__ import annotations

import io
from pathlib import Path
from typing import Iterable
from uuid import uuid4

from openpyxl import load_workbook
from pydantic import BaseModel
from pydantic_ai import BinaryContent
from pypdf import PdfReader

from email_automation.config import Settings
from email_automation.schemas import AttachmentDocument
from email_automation.services.openrouter import OpenRouterClient


class AttachmentService:
    def __init__(self, settings: Settings, openrouter: OpenRouterClient | None = None):
        self.settings = settings
        self.openrouter = openrouter

    async def process(
        self, attachments: Iterable[AttachmentDocument]
    ) -> list[AttachmentDocument]:
        processed: list[AttachmentDocument] = []
        for attachment in attachments:
            attachment.storage_path = await self._persist(attachment)
            attachment.extracted_text = await self._extract_text(attachment)
            processed.append(attachment)
        return processed

    async def _persist(self, attachment: AttachmentDocument) -> str:
        suffix = Path(attachment.filename).suffix
        target = self.settings.attachment_storage_path / f"{uuid4().hex}{suffix}"
        target.write_bytes(attachment.content_bytes)
        return str(target)

    async def _extract_text(self, attachment: AttachmentDocument) -> str:
        filename = attachment.filename.lower()
        if filename.endswith(".pdf"):
            return self._extract_pdf_text(attachment.content_bytes)
        if filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            return self._extract_workbook_text(attachment.content_bytes)
        if filename.endswith((".csv", ".txt")):
            return attachment.content_bytes.decode("utf-8", errors="ignore")
        if attachment.content_type and attachment.content_type.startswith("image/"):
            return await self._describe_image(attachment)
        return f"Attachment present but not parsed automatically: {attachment.filename} ({attachment.content_type or 'unknown'})"

    def _extract_pdf_text(self, content_bytes: bytes) -> str:
        reader = PdfReader(io.BytesIO(content_bytes))
        pages = []
        for page in reader.pages:
            pages.append(page.extract_text() or "")
        return "\n\n".join(filter(None, pages)).strip()

    def _extract_workbook_text(self, content_bytes: bytes) -> str:
        workbook = load_workbook(io.BytesIO(content_bytes), data_only=True)
        rows: list[str] = []
        for sheet in workbook.worksheets:
            rows.append(f"[Sheet: {sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                values = [
                    str(value).strip() for value in row if value not in (None, "")
                ]
                if values:
                    rows.append(" | ".join(values))
        return "\n".join(rows).strip()

    async def _describe_image(self, attachment: AttachmentDocument) -> str:
        if not self.openrouter or not attachment.content_type:
            return f"Image attachment detected: {attachment.filename}"

        content = [
            "Describe any products, model names, prices, quantities, and specs visible in this image. Return plain text.",
            BinaryContent(
                data=attachment.content_bytes, media_type=attachment.content_type
            ),
        ]
        try:
            result = await self.openrouter.complete_json(
                system_prompt=(
                    "You are extracting structured sales information from an attachment image. "
                    "Return JSON with a single key named summary."
                ),
                user_content=content,
                response_model=ImageSummary,
            )
        except Exception:
            return f"Image attachment detected: {attachment.filename}"
        image_summary = ImageSummary.model_validate(result)
        return image_summary.summary


class ImageSummary(BaseModel):
    summary: str
